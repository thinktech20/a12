"""Generate a simple .xlsx version of the Promotion Readiness Reflection + Plan drafts.

Companion to:
  - 01_Promotion-Readiness-Reflection_DRAFT.md
  - 02_Promotion-Plan_DRAFT.md

This is "Option A" — a plain xlsx with the content laid out in readable columns.
It does NOT replicate the manager's template formulas / Harvey Balls / dropdowns.
For that, manually paste cells into the official template at your own pace.
"""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

OUT = "/home/u560060992/dbx/ai-arch/promo/output-docs/Promo-Reflection-and-Plan_DRAFT.xlsx"

# ---- styles ----
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SECTION_FILL = PatternFill("solid", fgColor="D9E1F2")
SECTION_FONT = Font(bold=True, size=12)
COMP_FILL = PatternFill("solid", fgColor="FFF2CC")
COMP_FONT = Font(bold=True, size=11)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(wrap_text=True, vertical="center", horizontal="center")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header(cell):
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = CENTER
    cell.border = BORDER


def style_section(cell):
    cell.fill = SECTION_FILL
    cell.font = SECTION_FONT
    cell.alignment = WRAP
    cell.border = BORDER


def style_comp(cell):
    cell.fill = COMP_FILL
    cell.font = COMP_FONT
    cell.alignment = WRAP
    cell.border = BORDER


def style_body(cell):
    cell.alignment = WRAP
    cell.border = BORDER


wb = Workbook()

# =========================================================================
# Sheet 1: README
# =========================================================================
ws = wb.active
ws.title = "README"
ws.column_dimensions["A"].width = 120

readme_lines = [
    ("Promotion Readiness — DRAFT (v0)", True),
    ("", False),
    ("Candidate: Madhurima Saxena", False),
    ("Target title: Sr. Solutions Architect — Data & AI (Principal track)", False),
    ("Draft date: 2026-05-12", False),
    ("", False),
    ("Status: EARLY DRAFT for People Leader review.", True),
    ("", False),
    ("This workbook is a simplified version of the official Slalom template.", False),
    ("It does NOT include the Harvey Ball formulas, dropdowns, or hidden expectation columns.", False),
    ("To produce the final submission, paste the cell contents into the official template:", False),
    ("  - Reflection: Template - Promotion Readiness Reflection.xlsx → 'Reflection - To Principal' sheet", False),
    ("  - Plan: Template - Promotion Plan.pptx", False),
    ("", False),
    ("Sheets in this workbook:", True),
    ("  1. README — this sheet", False),
    ("  2. Reflection - Why & Stories — Section 1 (Why) + Section 2 (Anchor stories)", False),
    ("  3. Reflection - Competencies — Section 3, all 13 competency rows", False),
    ("  4. Reflection - Quantitative — Section 4 (utilization / revenue / role mastery)", False),
    ("  5. Reflection - Goals & Asks — Section 5 (goals + support needs) + Section 6 (asks for PL)", False),
    ("  6. Promotion Plan — 4-row plan table (one row per core competency)", False),
    ("", False),
    ("Status legend used in cells: ✅ filled · 🟡 partial · ⏳ TO DO (need user/PL input)", False),
    ("", False),
    ("Companion docs:", True),
    ("  - output-docs/01_Promotion-Readiness-Reflection_DRAFT.md", False),
    ("  - output-docs/02_Promotion-Plan_DRAFT.md", False),
    ("  - evidence-bank.md (story bank)", False),
    ("  - extracted/previous appraisal__Saxena, Madhurima 2025 EOY.md (Manager Eval source)", False),
    ("  - extracted/previous-recommendation__Feedback_Received.md (peer / mgr feedback source)", False),
]

for i, (text, bold) in enumerate(readme_lines, start=1):
    c = ws.cell(row=i, column=1, value=text)
    c.alignment = WRAP
    if bold:
        c.font = Font(bold=True, size=12)

print(f"README sheet written ({len(readme_lines)} rows).")

# =========================================================================
# Sheet 2: Why & Stories
# =========================================================================
ws = wb.create_sheet("Reflection - Why & Stories")
ws.column_dimensions["A"].width = 30
ws.column_dimensions["B"].width = 120

rows = [
    ("Section 1 — My Why", "", "section"),
    (
        "What's my why?",
        "I want to operate at the level where I'm not just delivering an architecture but owning the outcome — the trust, the team, the business case, and the recommendation that holds up in front of the customer's executives and our internal leadership. The Principal title formalizes what I've been operating as on Salesforce GIC and GE Vernova: workstream owner who can flex between deep technical execution, advisory, and pre-sales.",
        "body",
    ),
    (
        "What are my career desires?",
        "Continue as a customer-facing Solutions Architect in the Data & AI space — anchor at Slalom on Databricks + AWS + GenAI agentic patterns, lead pursuits end-to-end, and grow the next layer of Slalom architects through mentoring and the Databricks COE.",
        "body",
    ),
    (
        "How am I performing at current job expectations?",
        'Per 2025 EOY appraisal (Manager Evaluation): rated Above Expectations. Manager noted "her impact to multiply beyond the remit typical of a Senior Consultant" and "consistently demonstrated leader-like qualities… beyond those of a more specialized data engineer." Operating informally at Principal scope on Salesforce GIC (workstream lead) and GE Vernova (Slalom delivery owner for FSR pipeline).',
        "body",
    ),
    ("⏳ TO DO", "Tighten the 'why' into 2 sentences after PL conversation.", "body"),
    ("", "", "blank"),
    ("Section 2 — Anchor Stories", "", "section"),
    (
        "1. GE Vernova / AWS — SDG / FSR pipeline (2026, current) — HEADLINE",
        "Picked up the FSR data pipeline end-to-end on short notice when the GE architect raised to the PM that the team did not have the Databricks skillset, and the upstream-data gap had put production at risk. Delivered the production backfill (17,967 / 18,094 docs = 99.3%, 1.09M chunks, 0.43% failure rate, all corrupt-source PDFs). Now active design partner on the AdHoc Q&A Agent (Strands ReAct + Claude 4.6 Sonnet + MCP tools) and Unit Risk Agent.",
        "body",
    ),
    (
        "2. Salesforce GIC / ICM (2025–present) — PRIMARY DELIVERY",
        "Workstream Lead on multi-quarter analytics modernization. Surfaced ~33% of upstream data wasn't SOX-compliant. Authored the North Star Architecture, presented two future-state options to C-level + enterprise architects, drove ~80% reduction in manual effort. Engagement extended; client requested me back. NorCal H2 Mogul Award nomination.",
        "body",
    ),
    (
        "3. Apple / FruitCo (2023–2024) — DEEP TECH + MENTORING",
        "Sole Slalom architect ~12 months. Lakehouse on AWS Glue + PySpark + Iceberg + Airflow + Trino + Tableau. 90% data-accuracy improvement, customer adopted the architecture for all datasets (scope expansion), Customer Impact Award from Debra Ameerally (Apple).",
        "body",
    ),
    (
        "4. Pre-sales + Internal Slalom (cross-year) — GROW SLALOM",
        "BayREN RFP (~$2–3M D&A pipeline contribution per IW), Salesforce–Informatica Merger RFP, Genentech gRED, Proofpoint, MTC RFP, AWS MAP funding. Internal: Data Arch-Hive (May 2026), Databricks COE Insurance 360, weekly Databricks L&L sessions, Databricks Data Engineer Associate cert (2025).",
        "body",
    ),
]

for i, (label, text, kind) in enumerate(rows, start=1):
    a = ws.cell(row=i, column=1, value=label)
    b = ws.cell(row=i, column=2, value=text)
    if kind == "section":
        style_section(a)
        style_section(b)
    else:
        style_body(a)
        style_body(b)
    ws.row_dimensions[i].height = 60 if kind == "body" else 22

print(f"Why & Stories sheet written ({len(rows)} rows).")

# =========================================================================
# Sheet 3: Competencies (built up incrementally below)
# =========================================================================
ws = wb.create_sheet("Reflection - Competencies")
ws.column_dimensions["A"].width = 22
ws.column_dimensions["B"].width = 28
ws.column_dimensions["C"].width = 60
ws.column_dimensions["D"].width = 65
ws.column_dimensions["E"].width = 35

# Header row
headers = ["Core Competency", "Sub-Competency", "Principal Expectation", "Self Reflection — Examples / Feedback", "Self Rating + PL Input"]
for col, h in enumerate(headers, start=1):
    c = ws.cell(row=1, column=col, value=h)
    style_header(c)
ws.row_dimensions[1].height = 30

competency_rows = []  # accumulator; chunks below append to this

# ---- chunk 1: Deliver Exceptionally ----
competency_rows += [
    (
        "Deliver Exceptionally",
        "Execute Effectively",
        "Responsible for solution and/or overall project outcomes · Credits others publicly · Proactively removes obstacles (self & others) · Ensures multiple POVs are integrated · Creates clarity of roles and functions.",
        "• GEV / FSR (2026): Took end-to-end ownership of the production FSR data pipeline when the GE architect declined the Databricks scope to the PM. Drove backfill to 99.3% success (17,967 / 18,094 docs) in the planned 2–2.5 day window; hardened the run-loop (claim-based locking, retry caps, P1 telemetry, 41-check DQ validation). Owned the outcome, not just the deliverable.\n"
        '• Salesforce GIC (2025): Operated as workstream lead on a multi-quarter ICM modernization. SD Singh: "You have demonstrated leadership without authority… would have done a reasonably good job of being the engagement lead." [FB-SDSingh-2025-06-12]\n'
        '• 2025 EOY Manager Evaluation: "award-winning client delivery, and demonstrated ability to flex between data engineering, architectural, and stakeholder management responsibilities." [EOY25-MgrEval]',
        "Self: 4 — Consistently Demonstrates\nPL Input: ___",
    ),
    (
        "Deliver Exceptionally",
        "Develop High-Impact Solutions",
        'Applies best practices to accomplish project and organizational objectives · Enlists expertise to determine "best fit" solutions · Guides work (self & others) for prioritization, quality, accuracy.',
        "• GEV / FSR: Designed the run-loop so backfill and incremental run on the same code path; introduced claim-based parallel chunking, lifetime retry gate, run-log telemetry, and a 41-check standalone DQ job for SRE handoff.\n"
        "• Salesforce GIC: Designed two future-state architecture options (S3 + Snowflake vs S3 Lakehouse) and authored the North Star Architecture aligned with the client's enterprise platform direction. Presented to C-level + enterprise architects.\n"
        "• Apple / FruitCo: Lakehouse build on AWS Glue + PySpark + Iceberg + Airflow + Trino + Tableau, medallion architecture, Great Expectations for DQ. 90% data-accuracy improvement, customer adopted for all datasets.",
        "Self: 4 — Consistently Demonstrates\nPL Input: ___",
    ),
    (
        "Deliver Exceptionally",
        "Know & Serve Our Customers",
        "Creates authentic, trusting relationships at many levels · Understands informal org structures and constraints · Displays executive presence · Navigates discrepancies in customer and Slalom needs.",
        '• Tim Lin (Salesforce client Manager): "I\'ve never seen such data in all my years at Salesforce." [MID25]\n'
        '• Hong Wu (Salesforce client Director): "Will Madhurima have insight into it? Will she provide input into what that plan looks like?" — pulled into senior-level data-governance planning. [FB-SDSingh-2025-06-12]\n'
        '• Debra Ameerally (Apple senior exec): Customer Impact Award + "stands out for her quiet and diligent work." [FB-Brad-2024-04-10]\n'
        '• Maya Sandler: "Her ability to build a strong, trusting relationship with the client was particularly evident when it directly led to the project\'s extension for the next phase." [FB-MayaSandler-2025-06-20]\n'
        "• Growth area I own: tailoring communication for executive audiences — succinct, visually-engaging stories that 'connect the dots' for execs (called out by Matt Ngai, Maxine Leu, SD Singh, EOY25-MgrEval).\n"
        "• ⏳ Awaiting written feedback from Sreedhar Sannidhanam (GEV).",
        "Self: 3 — Sometimes Demonstrates\nPL Input: ___",
    ),
]

print(f"Competency rows so far: {len(competency_rows)}")

# ---- chunk 2: Grow Expertise ----
competency_rows += [
    (
        "Grow Expertise",
        "Deepen & Broaden Expertise",
        "Deepens and/or broadens in Functional / Technical roles · Recognized for expertise · Mentors others · Builds mastery in Pursuit & Engagement roles; increases exposure to Business roles.",
        '• Data Architect — Advanced+ (peer + manager validated). Sean Ipakchi: "Resident NorCal AWS Data Architecture expert." [FB-SeanIpakchi-2023-12-19]\n'
        "• Active broadening into GenAI / agentic patterns: Strands ReAct framework, Claude 4.6 Sonnet via FastAPI / Streamable HTTP, MCP-exposed data services, S3SessionManager, persona-based tools — solutioning + design partner on the GEV AdHoc Q&A Agent and Unit Risk Agent.\n"
        "• Databricks Data Engineer Associate certification (2025); AWS Certified Data Analytics – Specialty (2021); active across Databricks Lakehouse + Vector Search + Unity Catalog.\n"
        "• Mentoring — Erika Morris (Apple/FruitCo, end-to-end DE upskilling), Maya Sandler (AWS pipelines), Slalom colleagues for Databricks + AWS certifications.",
        "Self: 4 — Consistently Demonstrates\nPL Input: ___",
    ),
    (
        "Grow Expertise",
        "Broaden Slalom Knowledge",
        "Expands knowledge of Slalom business · Ensures Slalom stories are captured timely and accurately.",
        '• Data Arch-Hive session (May 2026) — co-presented FruitCo end-to-end Lakehouse + AI walkthrough with Raghav. Karthik Mannava: "the back-and-forth with the audience turned it into a genuine learning conversation… everyone walked away having learned something new." [FB-Karthik-2026-05-08]\n'
        "• Databricks COE — Insurance 360 — supported in-house solution build; dashboards showcased to the Databricks team and at the London Databricks meetup. [FB-Pragathi-2026-01-26]\n"
        "• Growth: more proactive Slalom-story capture (account anecdotes → reusable internal artifacts).",
        "Self: 3 — Sometimes Demonstrates\nPL Input: ___",
    ),
    (
        "Grow Expertise",
        "Build Slalom Capabilities",
        "May lend expertise to enhance capabilities · May contribute to cross-market community.",
        '• Databricks COE — weekly Databricks Learning sessions + Insurance 360 in-house solution. Emanuel Ward: "contributes to the Databricks Learning sessions on a weekly basis… key to supporting and growing our partnership with Databricks." [FB-EmanuelWard-2025-11-12]\n'
        "• Cross-market reach: dashboards from Insurance 360 showcased at the London Databricks meetup (cross-geo).\n"
        "• GenAI Cohort capstone (2024) — solution architect on the DMV-prep GenAI assistant.",
        "Self: 3 — Sometimes Demonstrates\nPL Input: ___",
    ),
]

print(f"Competency rows so far: {len(competency_rows)}")

# ---- chunk 3: Grow Slalom ----
competency_rows += [
    (
        "Grow Slalom",
        "Build Community",
        "Intentionally grows and manages professional network in delivery environments and/or area of expertise · Encourages others to engage in Slalom events · Models Slalom brand stewardship.",
        "• Active in NorCal AWS + Databricks community across delivery and L&L.\n"
        "• Networks built and maintained across: Salesforce GIC (Hongwu Wang, Tim Lin, SD Singh, Maya, Matt Ngai), Apple/FruitCo (Debra Ameerally, Pete Obringer, Julie Thomson, Erika), GEV (Sreedhar, Pranesh, Abhijeet, Tao), Slalom Databricks COE (Pragathi, Emanuel, Karthik), cross-eng (Maxine, Raphael, Aprajita).\n"
        "• Ongoing engagement with Slalom Databricks partnership team and AWS alliance team.\n"
        "• Growth: more outward Slalom brand stewardship (blog posts, public speaking, external community).",
        "Self: 3 — Sometimes Demonstrates\nPL Input: ___",
    ),
    (
        "Grow Slalom",
        "Manage Business Operations",
        "Proactively maximizes staffability (self & others) · Responsible for aspects of project financial & operational health (margin, managed revenue, staffing, engagement).",
        "• Utilization: 86.3% (2025), 93% (H1 2025). Above the 80%+ Principal bar.\n"
        '• Staffing leadership on Salesforce GIC — Maxine Leu: "Acted with high urgency to screen several Slalom candidates for a backfill / team expansion staffing need and is taking the lead to onboard the new Slalom team members." [FB-MaxineLeu-2025-06-11]\n'
        "• Operating as the Slalom delivery owner for the FSR data pipeline workstream on GEV (engagement health partnership with Sreedhar).\n"
        "• ⏳ Growth / TO DO: confirm managed-revenue numbers for Salesforce GIC and GEV with PL + account leads.",
        "Self: 3 — Sometimes Demonstrates\nPL Input: ___",
    ),
    (
        "Grow Slalom",
        "Drive Sales",
        "Proactively identifies opportunities for Slalom in delivery environments and/or area of expertise · Begins leading and/or solutioning for pursuits.",
        "• ~$2–3M in D&A pipeline influenced through pre-sales (per IW.pdf, anchored on BayREN RFP — technical solution lead, advanced Slalom to final selection).\n"
        "• Pre-sales contributions across: Salesforce–Informatica Merger RFP, Genentech gRED Sales, California Community Colleges, Proofpoint (data-ingestion review on short notice), MTC RFP, AWS MAP funding (zero-cost POC for healthcare client).\n"
        '• Raphael Vannson (Proofpoint pre-sales): "Madhurima quickly pivoted and helped reviewing the data ingestion approach… Slalom was able to make a confident and relevant presentation to the client and credentialize us on the data ingestion portion in addition to the AI portion of the solution." [FB-Raphael-2025-06-05]\n'
        "• Account extensions contributed to: Apple/FruitCo (architecture adopted for all datasets → scope expansion), Salesforce GIC (engagement extended on client request), Apple (engagement extended through end-of-FY into next FY).\n"
        "• ⏳ Honest gap: not yet led a pursuit end-to-end as Pursuit Lead — anchoring as a Promotion Plan goal.",
        "Self: 3 — Sometimes Demonstrates (approaching)\nPL Input: ___",
    ),
]

print(f"Competency rows so far: {len(competency_rows)}")

# ---- chunk 4: Lead ----
competency_rows += [
    (
        "Lead",
        "Lead & Develop Self",
        "Builds skills to identify and mitigate personal biases · Proactively manages self to protect energy, drive efficiency · Experiments with new approaches; fails fast and applies learnings · Openly and authentically shares values, assumptions, imperfections.",
        "• Stretched into GenAI / agentic patterns in 2025–2026 from a data-engineering / lakehouse foundation — Strands ReAct, MCP, Vector Search, embeddings + chunking strategies — and brought that into delivery on GEV.\n"
        "• Continuous learning: Databricks DE Associate (2025), AWS Data Analytics Specialty (2021), GenAI Cohort capstone (2024), weekly Databricks L&L participation.\n"
        '• 2025 EOY Manager Evaluation: highly engaged on professional development as an ongoing topic in 1:1s; "interest and time investment has been apparent and demonstrable." [EOY25-MgrEval]',
        "Self: 4 — Consistently Demonstrates\nPL Input: ___",
    ),
    (
        "Lead",
        "Lead & Develop Others",
        "Celebrates uniqueness and empowers others · Models the habit of regularly requesting and providing feedback · Encourages conversations that empower open dialogue · Supports others in defining their path forward.",
        '• Erika Morris (Apple/FruitCo mentee): "Madhurima has devoted time and effort to upskill me in data engineering by being extremely available for trouble-shooting, answering questions, and pair-programming whenever I encounter issues. She has been patient and understanding at all times, reserving judgement and making me feel very comfortable asking any questions I may have." + "This demonstrated Madhurima\'s potential as a future people leader." [FB-Erika-2024-07-01, FB-Erika-2024-11-22]\n'
        '• Maya Sandler: "Is a true leader in the AWS community in Slalom, inspiring and mentoring others." [FB-MayaSandler-2024-10-16]\n'
        "• Salesforce GIC — onboarded new Slalom team members during backfill / team expansion (per Maxine Leu feedback).\n"
        '• Nabor Reyna (iRhythm): "Empowered team members to ask questions, and grow their expertise. When teammates were confused she helped lead working sessions and helped remove obstacles." [FB-Nabor-2023-11-28]\n'
        "• Honest: no formal direct reports; mentoring is ad-hoc rather than structured.",
        "Self: 3 — Sometimes Demonstrates\nPL Input: ___",
    ),
    (
        "Lead",
        "Lead & Develop Teams",
        "Anticipates risks and proactively addresses obstacles to unblock the team, delegating as needed · Proactively removes barriers for productive conflict · Co-creates team vision, strategy, targeted outcomes · Establishes norms for open-mindedness and non-judgement.",
        "• GEV / FSR: unblocked the team by taking the Databricks pipeline scope when it had no owner; designed the run-loop so the team can operate it without me in the critical path (telemetry, DQ validation job, claim-based locking, retry caps).\n"
        '• Salesforce GIC: SD Singh: "That is how Slalom is perceived — we come in, make a significant impact, and leave the client better than where they started. You are doing that." [FB-SDSingh-2025-06-12]\n'
        '• 2025 EOY Manager Evaluation: "consistently demonstrated leader-like qualities within her delivery teams, and the adaptability to support both engineering and architectural outcomes, beyond those of a more specialized data engineer." [EOY25-MgrEval]\n'
        "• Growth: more proactive risk flagging to the broader team (not just 1:1s) and diplomatically challenging clients earlier when expectations are misaligned.",
        "Self: 3 — Sometimes Demonstrates\nPL Input: ___",
    ),
    (
        "Lead",
        "Serve As Formal People Leader (Optional)",
        "Optional w/ mastery & business need. Proactively connects employees to development experiences · Defines how directs' work connects to Slalom's goals · Ensures performance expectations + feedback · Translates org change.",
        "• No formal direct reports.\n"
        "• Informal mentoring captured under Lead & Develop Others.\n"
        "• Open to discussing whether to pursue formal PL responsibilities post-promotion.",
        "Self: 0 — Have not Started (Optional row)\nPL Input: ___",
    ),
]

print(f"Competency rows total: {len(competency_rows)}")

# (writing competency rows is below the chunks — keep moving)

# ---- write all competency rows to the sheet ----
for i, (core, sub, exp, ex, rating) in enumerate(competency_rows, start=2):
    a = ws.cell(row=i, column=1, value=core)
    b = ws.cell(row=i, column=2, value=sub)
    c = ws.cell(row=i, column=3, value=exp)
    d = ws.cell(row=i, column=4, value=ex)
    e = ws.cell(row=i, column=5, value=rating)
    style_comp(a)
    style_comp(b)
    style_body(c)
    style_body(d)
    style_body(e)
    ws.row_dimensions[i].height = 180

# =========================================================================
# Sheet 4: Quantitative
# =========================================================================
ws = wb.create_sheet("Reflection - Quantitative")
ws.column_dimensions["A"].width = 40
ws.column_dimensions["B"].width = 25
ws.column_dimensions["C"].width = 70
ws.column_dimensions["D"].width = 25
ws.column_dimensions["E"].width = 50

q_headers = ["Item", "Principal Bar", "My Position", "Status", "Note"]
for col, h in enumerate(q_headers, start=1):
    c = ws.cell(row=1, column=col, value=h)
    style_header(c)
ws.row_dimensions[1].height = 30

q_rows = [
    (
        "Utilization (Self)",
        "80%+",
        "86.3% (2025), 93% (H1)",
        "✅ Meets",
        "Above bar consistently.",
    ),
    (
        "Revenue (Primarily Sold or Managed)",
        "$1M",
        "Salesforce GIC managed revenue (multi-quarter, workstream lead) + GEV managed revenue (current) + ~$2–3M influenced pre-sales pipeline (BayREN-anchored, per IW)",
        "🟡 Partial — need numbers",
        "⏳ TO DO: confirm $ figures with SD Singh (Salesforce), Sreedhar Sannidhanam (GEV), Pete Obringer (FruitCo). Frame as managed revenue + influenced pipeline. Gap = leading a sold pursuit end-to-end.",
    ),
    (
        "1+ Functional / Technical Roles: Advanced+",
        "Advanced+",
        "Data Architect — Advanced+ (peer + manager validated); broadening into GenAI / agentic patterns",
        "✅ Meets",
        "",
    ),
    (
        "Engagement Lead or Delivery Solution Lead: Foundational+",
        "Foundational+",
        "Operating as workstream lead on Salesforce GIC; Slalom delivery owner for FSR pipeline on GEV",
        "🟡 Meets (informal)",
        'SD Singh: "would have done a reasonably good job of being the engagement lead." Open question: formalize the title.',
    ),
    (
        "Pursuit Lead or Sales Solution Lead: Foundational+",
        "Foundational+",
        "Contributor on BayREN, Informatica RFP, iRhythm POC, Proofpoint, MTC RFP, Cal Comm Colleges, Genentech",
        "🟡 Approaching",
        "Gap = lead a pursuit end-to-end as Pursuit Lead — anchoring this as a Promotion Plan goal.",
    ),
    (
        "People Leader: Proficient+",
        "Optional",
        "No formal directs; informal mentoring (Erika, Maya, Slalom colleagues for certs)",
        "0",
        "Optional at Principal — not a blocker.",
    ),
]

for i, row in enumerate(q_rows, start=2):
    for col, val in enumerate(row, start=1):
        c = ws.cell(row=i, column=col, value=val)
        style_body(c)
    ws.row_dimensions[i].height = 70

print(f"Quantitative sheet written ({len(q_rows)} rows).")

# =========================================================================
# Sheet 5: Goals & Asks
# =========================================================================
ws = wb.create_sheet("Reflection - Goals & Asks")
ws.column_dimensions["A"].width = 45
ws.column_dimensions["B"].width = 60
ws.column_dimensions["C"].width = 50

# Section header for Goals
c = ws.cell(row=1, column=1, value="Section 5 — Goals + Support Needs")
c.value = "Section 5 — Goals + Support Needs (these feed into the Promotion Plan)"
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
style_section(c)
ws.row_dimensions[1].height = 25

g_headers = ["Opportunity Area", "Goal to Address", "What I'd Need"]
for col, h in enumerate(g_headers, start=1):
    cc = ws.cell(row=2, column=col, value=h)
    style_header(cc)
ws.row_dimensions[2].height = 25

g_rows = [
    (
        "Visibility / Brand — Databricks + GenAI / Agents",
        "Build a sustained external-of-delivery presence so the practice associates Madhurima with Databricks + GenAI agentic patterns (not just AWS). PL flagged this as the #1 gap; Eugene's first question on promo will be 'what is she doing outside delivery?'.",
        "PL endorsement of topic + slot at Data Arch-Hive #2 (2026 Q4). 4–6 hours of bench / non-billable time per artifact. Co-presenter optional (Raghav model).",
    ),
    (
        "Pursuit Lead — lead a D&A / GenAI pursuit end-to-end (2026 H2)",
        "Take Pursuit Lead role on at least one D&A / GenAI pursuit in 2026 H2 — own the technical solution, write technical sections, partner on pricing, present to client. Closes the 'Drive Sales' Foundational+ gap.",
        "PL + Sales Lead to identify a fit-for-me pursuit. Brie's AI demo team is a likely on-ramp — follow up proactively. Pursuit-leadership / 'Selling the Slalom Way' training before taking the lead.",
    ),
    (
        "People-Leader Readiness — mentorship + practice fluency",
        "Build the muscle and visibility expected of a People Leader before formal reports arrive. Continue mentorship cadence (Maya + cert cohort), pick up one informal mentee through Databricks DE cert prep, get fluent with Salesforce-as-Slalom-uses-it (rates, margin, fixed-fee vs T&M, where to find numbers).",
        "Pointer / shadowing time with a Principal People Leader on the menial side (hours, vacation approvals, Salesforce internals). Mentee identified by PL or via cert cohort.",
    ),
]

for i, row in enumerate(g_rows, start=3):
    for col, val in enumerate(row, start=1):
        c = ws.cell(row=i, column=col, value=val)
        style_body(c)
    ws.row_dimensions[i].height = 70

# Section 6 header
sec6_row = 3 + len(g_rows) + 1
c = ws.cell(row=sec6_row, column=1, value="Section 6 — What I Need From PL (for our conversation)")
ws.merge_cells(start_row=sec6_row, start_column=1, end_row=sec6_row, end_column=3)
style_section(c)
ws.row_dimensions[sec6_row].height = 25

asks = [
    "1. Calibration — your read on each competency rating, especially Lead & Develop Teams + Drive Sales.",
    "2. Anchor confirmation — is the GEV / FSR pickup the right headline for Deliver Exceptionally + Lead?",
    "3. Pursuit Lead opportunity — help surface a fit-for-me D&A / GenAI pursuit in 2026 H2 to close the 'led a pursuit end-to-end' gap.",
    "4. Visibility / brand plays — endorse Data Arch-Hive #2 slot, Lunch & Learn demo cadence, internal article anchored on Eugene's 'AI in production' ask.",
    "5. Revenue numbers — guidance on Salesforce GIC + GEV managed-revenue figures, or who I should ask.",
    "6. Title framing — Sr. Solutions Architect — Data & AI vs. other Principal-level title options.",
]
for j, ask in enumerate(asks, start=sec6_row + 1):
    c = ws.cell(row=j, column=1, value=ask)
    ws.merge_cells(start_row=j, start_column=1, end_row=j, end_column=3)
    style_body(c)
    ws.row_dimensions[j].height = 35

# Open Items
oi_row = sec6_row + 1 + len(asks) + 1
c = ws.cell(row=oi_row, column=1, value="Open Items I'm Working On (after this conversation)")
ws.merge_cells(start_row=oi_row, start_column=1, end_row=oi_row, end_column=3)
style_section(c)
ws.row_dimensions[oi_row].height = 25

open_items = [
    "✅ PL conversation done May 13 — see promo/pl-conversation-2026-05-13.md. Headline: visibility/brand is the #1 gap; Eugene is gatekeeper.",
    "⏳ Follow up with Brie on the AI demo team (proactive ping — don't wait).",
    "⏳ Start AI-in-production artifact for Eugene (issues seen in prod, guardrails, team practices).",
    "⏳ Receive written feedback from Sreedhar Sannidhanam (GEV), Pranesh, Tao — drafts in feedback-asks.md.",
    "⏳ Confirm managed-revenue numbers (Salesforce GIC, GEV, FruitCo).",
    "⏳ Tighten 'why principal' to 2 sentences for the Reflection Summary slide.",
    "⏳ Schedule next checkpoint with PL.",
    "⏳ Move content into the official Slalom Reflection xlsx and Plan pptx templates.",
]
for j, oi in enumerate(open_items, start=oi_row + 1):
    c = ws.cell(row=j, column=1, value=oi)
    ws.merge_cells(start_row=j, start_column=1, end_row=j, end_column=3)
    style_body(c)
    ws.row_dimensions[j].height = 30

print(f"Goals & Asks sheet written.")

# =========================================================================
# Sheet 6: Promotion Plan
# =========================================================================
ws = wb.create_sheet("Promotion Plan")
ws.column_dimensions["A"].width = 35
ws.column_dimensions["B"].width = 50
ws.column_dimensions["C"].width = 50
ws.column_dimensions["D"].width = 60

p_headers = [
    "Focus Area (4 Core Competencies)",
    "'Done' Looks Like",
    "Goals (1–3)",
    "Activities (1–3)",
]
for col, h in enumerate(p_headers, start=1):
    c = ws.cell(row=1, column=col, value=h)
    style_header(c)
ws.row_dimensions[1].height = 35

plan_rows = [
    (
        "Grow Expertise + Grow Slalom — Visibility / Brand\n(Dev Team member: TBD — suggest Databricks COE lead Pragathi or Sounia Ghosh; co-presenter Raghav)",
        "Practice associates Madhurima with Databricks + GenAI / Agents (current brand still skews AWS-only). Sustained cadence of external-of-delivery contributions: Data Arch-Hive #2, recurring Lunch & Learn demos, 1 internal article on 'AI in production'. PL + Eugene see consistent visibility.",
        "1. Refresh brand from 'AWS expert' to 'Databricks + GenAI / Agents' inside the practice.\n2. Sustained, not one-off — at least 1 visible artifact per quarter.\n3. Anchor Eugene's stated ask: 'AI in production — issues, guardrails, team practices.'",
        "1. Data Arch-Hive #2 in 2026 Q3/Q4 — 'Production GenAI grounding pipeline + RAG pattern (FSR).'\n2. Brown-bag / Lunch & Learn demos: 'Build & productionize an agent on AWS', 'Build & productionize a RAG pipeline on Databricks'.\n3. Internal article (Slalom + Medium) on AI-in-production lessons + guardrails — anchors Eugene's ask.\n4. Engage in one Slalom employee community.",
    ),
    (
        "Grow Slalom — Drive Sales / Pursuit Lead\n(Dev Team member: TBD — a Slalom Sales Lead or SSL; possibly Brie for AI demo team on-ramp)",
        "Led one D&A / GenAI pursuit end-to-end as Pursuit Lead in 2026 H2 — owned technical solution, wrote technical sections, partnered on pricing, presented to client. Tracked outcome (won / lost / advanced). Comfort with Salesforce-as-Slalom-uses-it (rates, margin, fixed-fee vs T&M, where to find numbers).",
        "1. Take Pursuit Lead role on at least one pursuit in 2026 H2.\n2. Build comfort with the pricing + commercial side of pre-sales (not just the technical solution).\n3. Get fluent with Salesforce as the practice uses it (not as a tool).",
        "1. Proactively follow up with Brie on the AI demo team — likely on-ramp.\n2. Work with PL + Sales Lead to identify a fit-for-me D&A / GenAI pursuit.\n3. Complete 'Selling the Slalom Way' / pursuit-leadership training before taking the lead.\n4. Shadow one full pursuit cycle with an experienced Pursuit Lead.",
    ),
    (
        "Lead — People-Leader Readiness\n(Dev Team member: TBD — a Slalom Principal People Leader for shadowing on the menial side)",
        "Demonstrated People-Leader skills before formal reports arrive: at least one informal mentee guided through Databricks DE Associate cert prep; mentorship cadence with Maya + cert cohort sustained; familiarity with the menial side (hours, vacation approvals, Salesforce internals) so day-1 as People Leader is not a cold start.",
        "1. Continue and grow mentorship cadence (Maya, cert cohort, ad-hoc reach-outs).\n2. Take one informal mentee through Databricks DE Associate cert prep.\n3. Get exposure to the menial / operational side of being a People Leader before promo.",
        "1. Identify mentee through cert cohort or PL recommendation.\n2. Shadow a Slalom Principal on hours / vacation / Salesforce flows.\n3. Re-engage existing network (Khashin, Erica, Pete, Maxine, SD, Aprajita, Eugene, J) — not a one-off; sustained cadence so they vouch when promo time comes.",
    ),
    (
        "Deliver Exceptionally — Exec-Ready Communication\n(Dev Team member: TBD — Principal/Director with strong exec presence)",
        "Delivered 1–2 exec-level readouts (client C-level or Slalom LT) where the framing was visibly 'exec-ready' — succinct, visually engaging, business-outcome-led. PL + audience feedback confirms the shift.",
        "1. Build a personal pattern for 'exec-ready' deliverables.\n2. 'Connect the dots' for execs without losing the technical substance.\n3. Diplomatically surface risks / unrealistic expectations to clients earlier (not just 1:1s).",
        "1. Pair with a Slalom Principal/Director on 1–2 GEV or Salesforce GIC exec readouts in 2026 H2.\n2. Run drafts past PL before each exec readout for feedback.\n3. Self-rule: any risk in a 1:1 must surface in the next team / client touch-point with a recommendation.",
    ),
]

for i, row in enumerate(plan_rows, start=2):
    for col, val in enumerate(row, start=1):
        c = ws.cell(row=i, column=col, value=val)
        if col == 1:
            style_comp(c)
        else:
            style_body(c)
    ws.row_dimensions[i].height = 160

# Footer note
footer_row = 2 + len(plan_rows) + 1
c = ws.cell(row=footer_row, column=1, value="Note: Mirrors the official 'Template - Promotion Plan.pptx' structure. Final version to be moved into the pptx for submission.")
ws.merge_cells(start_row=footer_row, start_column=1, end_row=footer_row, end_column=4)
c.alignment = WRAP
c.font = Font(italic=True, size=10)

print("Promotion Plan sheet written.")

wb.save(OUT)
print(f"FINAL save to {OUT}")
print("Done.")
